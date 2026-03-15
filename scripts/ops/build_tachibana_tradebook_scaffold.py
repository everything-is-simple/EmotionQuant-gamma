from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


DEFAULT_OUTPUT_CSV = Path(
    "g:/EmotionQuant-gamma/normandy/03-execution/evidence/"
    "tachibana_tradebook_ledger_scaffold_1975_1976.csv"
)
DEFAULT_OUTPUT_JSON = Path(
    "g:/EmotionQuant-gamma/normandy/03-execution/evidence/"
    "tachibana_tradebook_scaffold_digest_20260315.json"
)


HOLIDAY_MAP = {
    "假": "假",
    "日": "日",
    "¼Ù": "假",
    "ÈÕ": "日",
}


def find_default_workbook() -> Path:
    candidates = [p for p in Path("G:/").rglob("*.xlsx") if "Pioneer" in p.name and "19751976" in p.name]
    if not candidates:
        raise FileNotFoundError("Could not find Tachibana Pioneer workbook under G:/")
    return candidates[-1]


def normalize_holiday(raw: object) -> str:
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text:
        return ""
    return HOLIDAY_MAP.get(text, text)


def build_rows(workbook_path: Path, symbol: str) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook.worksheets[0]

    rows: list[dict[str, object]] = []
    month_stats: dict[str, dict[str, int]] = {}

    for row_idx in range(1, sheet.max_row + 1):
        value = sheet.cell(row_idx, 1).value
        if not isinstance(value, datetime):
            continue

        holiday_tag = normalize_holiday(sheet.cell(row_idx, 2).value)
        close = sheet.cell(row_idx, 3).value
        trading_day = close is not None
        month_key = f"{value.year}-{value.month:02d}"

        month_stat = month_stats.setdefault(month_key, {"calendar_rows": 0, "close_rows": 0, "holiday_rows": 0})
        month_stat["calendar_rows"] += 1
        month_stat["close_rows"] += int(trading_day)
        month_stat["holiday_rows"] += int(bool(holiday_tag))

        rows.append(
            {
                "date": value.strftime("%Y-%m-%d"),
                "month": month_key,
                "symbol": symbol,
                "calendar_index": len(rows) + 1,
                "holiday_tag": holiday_tag,
                "is_trading_day": int(trading_day),
                "close": "" if close is None else int(close),
                "execution_price": "",
                "action": "",
                "buy_units": "",
                "sell_units": "",
                "open_units": "",
                "position_state": "",
                "reason_tag": "",
                "state_tag": "",
                "source": "xlsx_price_table_scaffold",
                "source_sheet": sheet.title,
                "source_row": row_idx,
                "trade_marker_readable": 0,
                "confidence": "A_price" if trading_day else "B_calendar",
                "note": "Trade markers are not machine-readable in workbook cells; manual confirmation required.",
            }
        )

    with ZipFile(workbook_path) as zf:
        names = zf.namelist()
        drawing_files = [n for n in names if n.startswith("xl/drawings/")]
        chart_files = [n for n in names if n.startswith("xl/charts/")]

    digest = {
        "source_workbook": str(workbook_path),
        "source_sheet": sheet.title,
        "rows_total": len(rows),
        "rows_with_close": sum(1 for row in rows if row["is_trading_day"] == 1),
        "rows_with_holiday_tag": sum(1 for row in rows if row["holiday_tag"]),
        "months_total": len(month_stats),
        "month_stats": month_stats,
        "machine_readable_trade_markers": False,
        "trade_marker_boundary": (
            "Price/calendar data is cell-readable, but trade and open-position markers are not "
            "present as machine-readable workbook cell values."
        ),
        "drawing_files": drawing_files,
        "chart_files": chart_files,
    }
    return rows, digest


def write_csv(rows: list[dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys())
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json(payload: dict[str, object], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the first Tachibana tradebook ledger scaffold.")
    parser.add_argument("--input-xlsx", type=Path, default=None, help="Optional workbook path.")
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV, help="CSV scaffold output path.")
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON, help="Digest JSON output path.")
    parser.add_argument("--symbol", default="PIONEER", help="Symbol label written into the scaffold.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.input_xlsx or find_default_workbook()
    rows, digest = build_rows(workbook_path=workbook_path, symbol=args.symbol)
    write_csv(rows=rows, output_path=args.output_csv)
    write_json(payload=digest, output_path=args.output_json)
    print(f"Workbook: {workbook_path}")
    print(f"Rows written: {len(rows)}")
    print(f"CSV: {args.output_csv}")
    print(f"JSON: {args.output_json}")


if __name__ == "__main__":
    main()
